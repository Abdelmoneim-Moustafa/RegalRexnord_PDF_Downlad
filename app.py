
"""
Smart Regal Rexnord PDF Downloader
===================================

Streamlit app for uploading an Excel file of product URLs, deduplicating them,
opening each page in a headless Chromium browser, finding the General
Specifications section, and saving the page as a PDF.

Improvements in this rewrite:
- cleaner Streamlit layout with upload preview and live run summary
- faster progress tracking using counters instead of repeated full rescans
- safer resume logic from a previous Master_Final.xlsx
- configurable save interval and browser restart behavior
- ZIP is built on disk instead of in memory for large batches
- more robust URL normalization and URL column detection
"""

from __future__ import annotations

import asyncio
import io
import random
import re
import shutil
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ───────────────────────────────────────────────────────────────
# Environment / Playwright setup
# ───────────────────────────────────────────────────────────────

if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

PLAYWRIGHT_OK = True
PLAYWRIGHT_IMPORT_ERROR = ""
try:
    from playwright.async_api import TimeoutError as PWTimeout
    from playwright.async_api import async_playwright
    from playwright_stealth import Stealth
except Exception as exc:  # noqa: BLE001
    PLAYWRIGHT_OK = False
    PLAYWRIGHT_IMPORT_ERROR = str(exc)


@st.cache_resource(show_spinner="Setting up the browser (first run only, ~30-60s)...")
def ensure_chromium_installed() -> tuple[bool, str]:
    """
    Download the Chromium browser binary if it isn't already present.

    `pip install playwright` only installs the Python automation library —
    the actual browser has to be fetched separately. Locally this is done
    once via `playwright install chromium` in a terminal. Hosted platforms
    like Streamlit Community Cloud have no terminal step between installing
    requirements.txt and starting the app, so this does the equivalent at
    runtime instead. @st.cache_resource means it only actually runs once
    per running container, not on every script rerun.
    """
    import subprocess

    result = subprocess.run(
        [sys.executable, "-m", "playwright", "install", "chromium"],
        capture_output=True, text=True, timeout=300,
    )
    if result.returncode != 0:
        return False, (result.stderr or result.stdout)[-2000:]
    return True, ""


if PLAYWRIGHT_OK:
    _chromium_ok, _chromium_err = ensure_chromium_installed()
    if not _chromium_ok:
        PLAYWRIGHT_OK = False
        PLAYWRIGHT_IMPORT_ERROR = (
            "Chromium installation failed.\n\n"
            "If you're on Streamlit Community Cloud, this is almost always "
            "a missing system library — make sure packages.txt is present "
            "at the repo root and the app has been rebuilt since adding it.\n\n"
            f"Raw error (last 2000 chars):\n{_chromium_err}"
        )

# ───────────────────────────────────────────────────────────────
# Constants
# ───────────────────────────────────────────────────────────────

PROJECT_DIR = Path("Regal Rexnord Corporation")
PDF_DIR = PROJECT_DIR / "PDFs"
ERR_DIR = PROJECT_DIR / "Errors"
LOG_DIR = PROJECT_DIR / "Logs"
MASTER_PATH = PROJECT_DIR / "Master_Final.xlsx"
PROFILE_DIR = PROJECT_DIR / ".browser_profile"

HOMEPAGE = "https://www.regalrexnord.com/"
RUN_DATE = datetime.now().strftime("%Y%m%d")
TODAY_DISPLAY = datetime.now().strftime("%Y-%m-%d")

BROWSER_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/137.0.0.0 Safari/537.36"
)

SPEC_SELECTORS = [
    'button:has-text("General Specifications")',
    'button:has-text("Specifications")',
    'a:has-text("General Specifications")',
    'a:has-text("Specifications")',
    '[role="tab"]:has-text("Specifications")',
    '[class*="tab"]:has-text("Specifications")',
    '[class*="accordion"]:has-text("Specifications")',
]

HEADERS = [
    "Row_ID",
    "URL",
    "Part_Number",
    "Duplicate_Flag",
    "Status",
    "PDF_File_Name",
    "Note",
]

STATUS_COLORS = {
    "Success": "C6EFCE",
    "Duplicate": "FFEB9C",
    "Failed": "FFC7CE",
    "Pending": "FFFFFF",
}

# ───────────────────────────────────────────────────────────────
# Helpers
# ───────────────────────────────────────────────────────────────

def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def normalize_url(url: str) -> str:
    """Normalize URLs for deduplication and resume matching."""
    u = _safe_text(url).lower()
    u = re.sub(r"#.*$", "", u)
    u = re.sub(r"\?.*$", "", u)
    u = u.rstrip("/")
    return u


def part_from_url(url: str) -> str:
    """Best-effort part number extraction from a URL slug."""
    slug = _safe_text(url).rstrip("/").split("/")[-1]
    slug = re.sub(r"\?.*$", "", slug)
    slug = re.sub(r"#.*$", "", slug)

    m = re.search(r"(\d{4}-\d{5})$", slug)
    if m:
        return m.group(1).upper()

    m = re.search(r"([A-Za-z]{1,6}\d{4,}-\d{2,})$", slug)
    if m:
        return m.group(1).upper()

    parts = [p for p in slug.split("-") if p]
    if not parts:
        return slug.upper()

    last = parts[-1]
    if (last.isdigit() and len(last) >= 4) or re.match(r"^[A-Za-z0-9]{3,}$", last):
        return last.upper()

    return ("-".join(parts[-2:]) if len(parts) >= 2 else slug).upper()


def detect_url_column(df: pd.DataFrame) -> str | None:
    candidates = [
        "COnlineUrl",
        "URL",
        "Url",
        "url",
        "Link",
        "link",
        "Product URL",
        "ProductURL",
    ]
    cols = [str(c).strip() for c in df.columns]
    mapping = dict(zip(cols, df.columns))

    for candidate in candidates:
        if candidate in mapping:
            return mapping[candidate]

    for col in df.columns:
        sample = df[col].dropna().astype(str).head(25)
        if any(v.strip().lower().startswith(("http://", "https://")) for v in sample):
            return col
    return None


def _header(ws, cols, color="003087"):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


def _autowidth(ws, min_w=10, max_w=80):
    for col in ws.columns:
        ltr = get_column_letter(col[0].column)
        length = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[ltr].width = max(min_w, min(max_w, length + 2))


def get_max_pdf_number() -> int:
    """Continue numbering from existing PDFs if the app is resumed."""
    max_n = 0
    if PDF_DIR.exists():
        for f in PDF_DIR.glob("PdfFile*.pdf"):
            m = re.match(r"PdfFile(\d{6})_", f.name)
            if m:
                max_n = max(max_n, int(m.group(1)))
    return max_n


def load_previous_results(path: Path) -> dict[str, dict[str, str]]:
    """Load previous Master_Final.xlsx for resume support."""
    result: dict[str, dict[str, str]] = {}
    if not path.exists():
        return result

    try:
        wb = load_workbook(path)
        ws = wb["Results"] if "Results" in wb.sheetnames else wb.active
        headers = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}
        required = {"URL", "Status", "PDF_File_Name", "Note"}
        if not required.issubset(idx.keys()):
            return result

        for row in ws.iter_rows(min_row=2, values_only=True):
            url = row[idx["URL"]]
            if not url:
                continue
            result[normalize_url(str(url))] = {
                "status": _safe_text(row[idx["Status"]]),
                "pdf_name": _safe_text(row[idx["PDF_File_Name"]]),
                "note": _safe_text(row[idx["Note"]]),
            }
    except Exception:
        return {}
    return result


def save_master(rows: list[dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    _header(ws, HEADERS)

    for r in rows:
        ws.append([
            r["row_id"],
            r["url"],
            r["part"],
            r["duplicate"],
            r["status"],
            r.get("pdf_name", "NULL"),
            r.get("note", ""),
        ])
        row_idx = ws.max_row
        fill_color = STATUS_COLORS.get(r["status"], "FFFFFF")
        for ci in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=ci)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    _autowidth(ws)
    ws.column_dimensions["B"].width = 80
    wb.save(path)


def save_errors(rows: list[dict[str, Any]], path: Path) -> None:
    failed = [r for r in rows if r["status"] == "Failed"]
    if not failed:
        if path.exists():
            path.unlink(missing_ok=True)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Failed URLs"
    _header(ws, ["Row_ID", "URL", "Error", "Timestamp"])

    for r in failed:
        ws.append([r["row_id"], r["url"], r.get("note", ""), TODAY_DISPLAY])

    _autowidth(ws)
    ws.column_dimensions["B"].width = 80
    wb.save(path)


def save_log(rows: list[dict[str, Any]], pdf_n: int, start: datetime, end: datetime, path: Path) -> None:
    total = len(rows)
    unique = sum(1 for r in rows if r["duplicate"] == "No")
    success = sum(1 for r in rows if r["status"] == "Success")
    failed = sum(1 for r in rows if r["status"] == "Failed")
    pending = sum(1 for r in rows if r["status"] == "Pending")
    dur = round((end - start).total_seconds(), 1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _header(ws, ["Metric", "Value"])

    metrics = [
        ("Total URLs", total),
        ("Unique", unique),
        ("Duplicates", total - unique),
        ("Successful", success),
        ("Failed", failed),
        ("Not yet processed", pending),
        ("PDFs in folder", pdf_n),
        ("Start Time", start.strftime("%Y-%m-%d %H:%M:%S")),
        ("End Time", end.strftime("%Y-%m-%d %H:%M:%S")),
        ("Duration (s)", dur),
        ("Duration (min)", round(dur / 60, 1)),
        ("Run Date", TODAY_DISPLAY),
    ]

    for k, v in metrics:
        ws.append([k, v])
        row_idx = ws.max_row
        ws.cell(row_idx, 1).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row_idx, 2).font = Font(name="Arial", size=10)
        ws.cell(row_idx, 2).alignment = Alignment(horizontal="right")

    _autowidth(ws)
    wb.save(path)


def build_zip_file(folder: Path) -> Path:
    """Create a ZIP on disk to avoid large in-memory blobs."""
    folder.parent.mkdir(parents=True, exist_ok=True)
    fd, zip_name = tempfile.mkstemp(prefix="regal_rexnord_", suffix=".zip")
    Path(zip_name).unlink(missing_ok=True)  # reopen via zipfile

    with zipfile.ZipFile(zip_name, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in folder.rglob("*"):
            if path.is_file() and ".browser_profile" not in path.parts:
                zf.write(path, path.relative_to(folder.parent))
    return Path(zip_name)


def _ensure_dirs() -> None:
    for d in (PROJECT_DIR, PDF_DIR, ERR_DIR, LOG_DIR, PROFILE_DIR):
        d.mkdir(parents=True, exist_ok=True)


def _extract_urls_from_df(df: pd.DataFrame, url_col: str) -> list[str]:
    urls: list[str] = []
    for value in df[url_col].tolist():
        text = _safe_text(value)
        if text:
            urls.append(text)
    return urls


def _download_link_text(path: Path) -> str:
    return f"{path.name} ({path.stat().st_size / 1024 / 1024:.1f} MB)"


# ───────────────────────────────────────────────────────────────
# Browser automation helpers
# ───────────────────────────────────────────────────────────────

async def launch_browser(pw):
    context = await pw.chromium.launch_persistent_context(
        user_data_dir=str(PROFILE_DIR),
        headless=True,
        user_agent=BROWSER_UA,
        locale="en-US",
        viewport={"width": 1400, "height": 1000},
        args=[
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--disable-dev-shm-usage",
            "--disable-blink-features=AutomationControlled",
        ],
    )
    page = context.pages[0] if context.pages else await context.new_page()

    try:
        await Stealth().apply_stealth_async(page)
    except Exception:
        pass

    try:
        await context.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
        )
    except Exception:
        pass

    return context, page


async def warmup(page, log):
    log("Warming up browser session...")
    try:
        await page.goto(HOMEPAGE, wait_until="domcontentloaded", timeout=30_000)
        await page.wait_for_timeout(2_000)
        log("Browser session ready.")
    except Exception as exc:  # noqa: BLE001
        log(f"Warmup warning: {exc}")


async def save_pdf(page, url: str, pdf_path: str) -> tuple[str, str, int]:
    """
    Returns:
      ("success" | "403" | "error", note, http_status)
    """
    try:
        response = await page.goto(url, wait_until="domcontentloaded", timeout=60_000, referer=HOMEPAGE)
        status = response.status if response else 0
    except PWTimeout:
        return "error", "Timeout", 0
    except Exception as exc:  # noqa: BLE001
        return "error", str(exc), 0

    if status == 403:
        return "403", "HTTP 403", 403
    if status >= 400:
        return "error", f"HTTP {status}", status

    try:
        await page.wait_for_load_state("networkidle", timeout=15_000)
    except Exception:
        pass

    await page.wait_for_timeout(1_500)

    clicked = False
    for sel in SPEC_SELECTORS:
        try:
            locator = page.locator(sel).first
            await locator.click(timeout=3_000)
            await page.wait_for_timeout(1_000)
            try:
                await page.wait_for_load_state("networkidle", timeout=10_000)
            except Exception:
                pass
            clicked = True
            break
        except Exception:
            continue

    try:
        out_path = Path(pdf_path)
        if out_path.exists():
            out_path.unlink(missing_ok=True)

        await page.pdf(
            path=str(out_path),
            format="Letter",
            print_background=True,
            margin={"top": "0.5in", "bottom": "0.5in", "left": "0.5in", "right": "0.5in"},
        )
        note = "specs button clicked" if clicked else "no specs button found"
        return "success", note, status
    except Exception as exc:  # noqa: BLE001
        try:
            Path(pdf_path).unlink(missing_ok=True)
        except Exception:
            pass
        return "error", f"PDF print failed: {exc}", status


async def restart_browser(pw, old_context, log, reason: str):
    log(reason)
    try:
        await old_context.close()
    except Exception:
        pass
    context, page = await launch_browser(pw)
    await warmup(page, log)
    return context, page


# ───────────────────────────────────────────────────────────────
# Processing engine
# ───────────────────────────────────────────────────────────────

async def run_processing(url_list: list[str], settings: dict[str, Any], ui: dict[str, Any]):
    start = datetime.now()
    _ensure_dirs()

    # Dedupe and initialize rows
    seen: dict[str, int] = {}
    rows: list[dict[str, Any]] = []

    for idx, raw in enumerate(url_list, 1):
        url = _safe_text(raw)
        norm = normalize_url(url)
        part = part_from_url(url)
        if not url:
            rows.append(
                dict(
                    row_id=idx,
                    url="",
                    norm="",
                    part=part,
                    duplicate="No",
                    status="Failed",
                    pdf_name="NULL",
                    note="Blank URL",
                )
            )
            continue

        if norm in seen:
            rows.append(
                dict(
                    row_id=idx,
                    url=url,
                    norm=norm,
                    part=part,
                    duplicate="Yes",
                    status="Duplicate",
                    pdf_name="NULL",
                    note="Duplicate URL",
                )
            )
        else:
            seen[norm] = idx
            rows.append(
                dict(
                    row_id=idx,
                    url=url,
                    norm=norm,
                    part=part,
                    duplicate="No",
                    status="Pending",
                    pdf_name="",
                    note="",
                )
            )

    total = len(rows)
    unique = sum(1 for r in rows if r["duplicate"] == "No")
    duplicates = total - unique

    previous = load_previous_results(MASTER_PATH)
    resumed = 0
    for r in rows:
        if r["duplicate"] == "No" and r["norm"] in previous:
            prev = previous[r["norm"]]
            if prev["status"] == "Success":
                r.update(status="Success", pdf_name=prev["pdf_name"], note=prev["note"])
                resumed += 1

    pdf_n = get_max_pdf_number()
    remaining = sum(1 for r in rows if r["duplicate"] == "No" and r["status"] == "Pending")
    success_n = sum(1 for r in rows if r["status"] == "Success")
    failed_n = sum(1 for r in rows if r["status"] == "Failed")

    ui["log"](
        f"Loaded {total} rows | Unique: {unique} | Duplicates: {duplicates} | "
        f"Resumed: {resumed} | Remaining: {remaining}"
    )
    ui["metrics"](success_n, failed_n, duplicates, remaining)
    ui["progress"](0 if unique == 0 else (unique - remaining) / unique)

    if remaining == 0:
        ui["log"]("Nothing left to process. All unique URLs are already completed.")
        end = datetime.now()
        save_master(rows, MASTER_PATH)
        save_errors(rows, ERR_DIR / "Failed_Invalid_URLs.xlsx")
        save_log(rows, pdf_n, start, end, LOG_DIR / "Processing_Log.xlsx")
        return dict(rows=rows, pdf_n=pdf_n, gave_up=False, start=start, end=end)

    consecutive_403 = 0
    gave_up = False
    processed_unique = 0
    last_save_at = 0

    async with async_playwright() as pw:
        context, page = await launch_browser(pw)
        await warmup(page, ui["log"])

        try:
            for row in rows:
                idx = row["row_id"]
                url = row["url"]

                if row["duplicate"] == "Yes":
                    continue
                if row["status"] == "Success":
                    continue

                if not re.match(r"^https?://", url, flags=re.IGNORECASE):
                    row.update(status="Failed", note="Invalid URL")
                    failed_n += 1
                    processed_unique += 1
                    ui["log"](f"[{idx}] INVALID URL")
                else:
                    next_pdf_n = pdf_n + 1
                    pdf_name = f"PdfFile{next_pdf_n:06d}_{RUN_DATE}_{row['part']}.pdf"
                    pdf_path = str(PDF_DIR / pdf_name)

                    result, note, _status_code = await save_pdf(page, url, pdf_path)

                    if result == "success":
                        pdf_n = next_pdf_n
                        row.update(status="Success", pdf_name=pdf_name, note=note)
                        success_n += 1
                        consecutive_403 = 0
                        ui["log"](f"[{idx}] OK -> {pdf_name} ({note})")
                    elif result == "403":
                        row.update(status="Failed", note=f"HTTP 403 (consecutive #{consecutive_403 + 1})")
                        failed_n += 1
                        consecutive_403 += 1
                        ui["log"](f"[{idx}] BLOCKED (HTTP 403) — {consecutive_403} in a row")
                    else:
                        row.update(status="Failed", note=note)
                        failed_n += 1
                        consecutive_403 = 0
                        ui["log"](f"[{idx}] FAILED — {note}")

                    processed_unique += 1

                remaining = sum(1 for r in rows if r["duplicate"] == "No" and r["status"] == "Pending")
                ui["metrics"](success_n, failed_n, duplicates, remaining)
                ui["progress"](0 if unique == 0 else (unique - remaining) / unique)

                if processed_unique - last_save_at >= settings["save_every"]:
                    save_master(rows, MASTER_PATH)
                    last_save_at = processed_unique

                if consecutive_403 == settings["soft_threshold"]:
                    await asyncio.sleep(settings["soft_seconds"])
                    context, page = await restart_browser(
                        pw,
                        context,
                        ui["log"],
                        f"{consecutive_403} consecutive blocks — pausing and refreshing browser session.",
                    )
                elif consecutive_403 == settings["hard_threshold"]:
                    await asyncio.sleep(settings["hard_seconds"])
                    context, page = await restart_browser(
                        pw,
                        context,
                        ui["log"],
                        f"{consecutive_403} consecutive blocks — long pause and browser refresh.",
                    )
                elif consecutive_403 >= settings["giveup_threshold"]:
                    ui["log"](
                        f"{consecutive_403} consecutive blocks — stopping safely. "
                        "Progress has been saved."
                    )
                    gave_up = True
                    break

                await asyncio.sleep(random.uniform(settings["delay_min"], settings["delay_max"]))

        finally:
            try:
                await context.close()
            except Exception:
                pass

    end = datetime.now()
    save_master(rows, MASTER_PATH)
    save_errors(rows, ERR_DIR / "Failed_Invalid_URLs.xlsx")
    save_log(rows, pdf_n, start, end, LOG_DIR / "Processing_Log.xlsx")

    return dict(rows=rows, pdf_n=pdf_n, gave_up=gave_up, start=start, end=end)


def run_async(coro):
    try:
        return asyncio.run(coro)
    except RuntimeError as exc:
        if "cannot be called from a running event loop" in str(exc):
            import nest_asyncio

            nest_asyncio.apply()
            loop = asyncio.get_event_loop()
            return loop.run_until_complete(coro)
        raise


# ───────────────────────────────────────────────────────────────
# Streamlit page
# ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Regal Rexnord PDF Downloader",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.title("🏭 Regal Rexnord — Smart Specifications PDF Downloader")
st.caption(
    "Upload an Excel file of product URLs, review a quick preview, and run a "
    "smart browser workflow that resumes automatically and saves progress as it goes."
)

if not PLAYWRIGHT_OK:
    st.error(
        "Playwright/Chromium is not ready.\n\n"
        "If you're running this **locally**, install the missing packages "
        "and Chromium in a terminal, then restart the app:\n\n"
        "```bash\n"
        "pip install -r requirements.txt\n"
        "playwright install chromium\n"
        "```\n\n"
        "If you're on **Streamlit Community Cloud**, make sure `packages.txt` "
        "exists at the repo root (it provides system libraries Chromium needs) "
        "and that the app has rebuilt since you added it — check 'Manage app' "
        "→ logs for the apt-get step.\n\n"
        f"Details:\n{PLAYWRIGHT_IMPORT_ERROR}"
    )
    st.stop()

with st.sidebar:
    st.header("⚙️ Run settings")
    delay_min, delay_max = st.slider(
        "Delay between requests (seconds)",
        3,
        30,
        (8, 15),
        help="Random delay between page loads.",
    )
    save_every = st.number_input(
        "Save progress every N processed URLs",
        min_value=1,
        max_value=500,
        value=25,
        step=5,
        help="Larger values are faster. Smaller values save more often.",
    )

    st.divider()
    st.subheader("Blocking response")
    soft_threshold = st.number_input("Short pause after N blocks", 1, 20, 3)
    soft_seconds = st.number_input("Short pause duration (sec)", 10, 1800, 60)
    hard_threshold = st.number_input("Long pause after N blocks", 1, 50, 6)
    hard_seconds = st.number_input("Long pause duration (sec)", 30, 3600, 300)
    giveup_threshold = st.number_input("Stop after N blocks", 2, 100, 10)

    st.divider()
    st.subheader("Storage")
    st.caption(f"Output folder: `{PROJECT_DIR.resolve()}`")
    st.caption("PDFs are saved directly on disk and also available in the download section.")

    st.divider()
    st.subheader("Reset")
    if st.button("Clear progress tracking (keep PDFs)", use_container_width=True):
        MASTER_PATH.unlink(missing_ok=True)
        st.success("Progress tracking removed. PDFs were kept.")
        st.rerun()

    wipe_confirm = st.checkbox("I understand — delete ALL output files")
    if st.button("Delete everything and start fresh", disabled=not wipe_confirm, use_container_width=True):
        shutil.rmtree(PROJECT_DIR, ignore_errors=True)
        st.success("All output files deleted.")
        st.rerun()

left, right = st.columns([1.45, 1.0], vertical_alignment="top")

with left:
    st.subheader("1) Upload")
    uploaded_file = st.file_uploader("Select Excel file", type=["xlsx", "xls"])

    if uploaded_file:
        try:
            df = pd.read_excel(uploaded_file, dtype=str)
            df.columns = df.columns.astype(str).str.strip()
        except Exception as exc:
            st.error(f"Could not read the Excel file: {exc}")
            st.stop()

        url_col = detect_url_column(df)
        st.write(f"Rows loaded: **{len(df):,}**")
        if url_col is None:
            st.error(f"No URL column found. Columns: {list(df.columns)}")
            st.stop()

        url_values = _extract_urls_from_df(df, url_col)
        unique_preview = len({normalize_url(u) for u in url_values if u})

        st.success(f'URL column detected: **{url_col}**  |  URLs found: **{len(url_values):,}**  |  Unique: **{unique_preview:,}**')

        with st.expander("Preview the file", expanded=True):
            st.dataframe(df.head(20), use_container_width=True, height=320)

        st.subheader("2) Start")
        start_clicked = st.button("🚀 Start Processing", type="primary", use_container_width=True)

        if start_clicked:
            _ensure_dirs()

            status_box = st.empty()
            progress_bar = st.progress(0.0)

            c1, c2, c3, c4 = st.columns(4)
            m_success = c1.metric("✅ Success", 0)
            m_failed = c2.metric("❌ Failed", 0)
            m_dupes = c3.metric("⚠️ Duplicates", 0)
            m_remaining = c4.metric("⏳ Remaining", 0)

            log_lines: list[str] = []
            log_box = st.empty()

            def log(msg: str):
                stamp = datetime.now().strftime("%H:%M:%S")
                log_lines.append(f"[{stamp}] {msg}")
                log_box.code("\n".join(log_lines[-250:]), language=None)

            def metrics(success, failed, dupes, remaining):
                m_success.metric("✅ Success", success)
                m_failed.metric("❌ Failed", failed)
                m_dupes.metric("⚠️ Duplicates", dupes)
                m_remaining.metric("⏳ Remaining", remaining)

            def progress(fraction):
                progress_bar.progress(min(max(float(fraction), 0.0), 1.0))

            ui = dict(log=log, metrics=metrics, progress=progress)
            settings = dict(
                delay_min=float(delay_min),
                delay_max=float(delay_max),
                save_every=int(save_every),
                soft_threshold=int(soft_threshold),
                soft_seconds=int(soft_seconds),
                hard_threshold=int(hard_threshold),
                hard_seconds=int(hard_seconds),
                giveup_threshold=int(giveup_threshold),
            )

            status_box.info("Processing started. The log and counters will update below.")
            result = run_async(run_processing(url_values, settings, ui))

            elapsed = round((result["end"] - result["start"]).total_seconds(), 1)
            if result["gave_up"]:
                status_box.warning(
                    f"Paused after repeated blocks. Progress is saved. "
                    f"Elapsed time: {elapsed}s."
                )
            else:
                status_box.success(f"Completed successfully. Elapsed time: {elapsed}s.")

            st.session_state["last_run_finished"] = True
            st.session_state["last_run_time"] = result["end"].strftime("%Y-%m-%d %H:%M:%S")
            st.rerun()
    else:
        st.info("Upload an Excel file to begin.")

with right:
    st.subheader("Smart summary")
    if MASTER_PATH.exists():
        prev = load_previous_results(MASTER_PATH)
        completed = sum(1 for v in prev.values() if v["status"] == "Success")
        failures = sum(1 for v in prev.values() if v["status"] == "Failed")
        st.metric("Previously completed", completed)
        st.metric("Previously failed", failures)
        st.caption("The app will skip URLs already marked Success.")
    else:
        st.info("No previous progress file found yet.")

    st.subheader("Files")
    st.caption(
        "⚠️ If this app is running on **Streamlit Community Cloud**, its disk "
        "is temporary — files can be lost on a redeploy or after the app "
        "sleeps from inactivity. Download anything you need before closing "
        "this tab. (Running locally, files persist normally on your disk.)"
    )
    if MASTER_PATH.exists():
        with open(MASTER_PATH, "rb") as f:
            st.download_button(
                "📄 Download Master_Final.xlsx",
                f,
                file_name="Master_Final.xlsx",
                use_container_width=True,
            )

    failed_path = ERR_DIR / "Failed_Invalid_URLs.xlsx"
    if failed_path.exists():
        with open(failed_path, "rb") as f:
            st.download_button(
                "⚠️ Download Failed_Invalid_URLs.xlsx",
                f,
                file_name="Failed_Invalid_URLs.xlsx",
                use_container_width=True,
            )

    pdf_count = len(list(PDF_DIR.glob("*.pdf"))) if PDF_DIR.exists() else 0
    st.caption(f"Current PDF count: **{pdf_count:,}**")

    if PROJECT_DIR.exists():
        st.divider()
        if st.button("📦 Build ZIP package", use_container_width=True):
            with st.spinner("Building ZIP package..."):
                zip_path = build_zip_file(PROJECT_DIR)
            with open(zip_path, "rb") as zf:
                st.download_button(
                    "⬇️ Download ZIP",
                    zf,
                    file_name=f"Regal_Rexnord_Output_{RUN_DATE}.zip",
                    mime="application/zip",
                    use_container_width=True,
                )
            try:
                zip_path.unlink(missing_ok=True)
            except Exception:
                pass

    st.divider()
    st.subheader("Notes")
    st.caption(
        "The browser runs in headless mode because PDF export requires it. "
        "Progress is written to disk as the app runs, which allows resuming "
        "within the same running session — but see the storage warning above "
        "if you're on a hosted/cloud deployment."
    )

if MASTER_PATH.exists():
    st.divider()
    st.subheader("Latest output summary")
    try:
        wb = load_workbook(MASTER_PATH, data_only=True)
        ws = wb["Results"] if "Results" in wb.sheetnames else wb.active
        df_master = pd.DataFrame(ws.values)
        if not df_master.empty:
            df_master.columns = df_master.iloc[0]
            df_master = df_master.iloc[1:].reset_index(drop=True)
            st.dataframe(df_master.tail(25), use_container_width=True, height=320)
    except Exception as exc:
        st.warning(f"Could not preview the summary file: {exc}")
